# -*- coding: utf-8 -*-
"""
Excel 处理工具模块
支持读写 Excel 文件，用于数据导入导出
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import List, Dict, Tuple
import os


class ExcelHandler:
    """Excel 文件处理器"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def create_workbook(self, title: str = "收支管理系统"):
        """创建新的工作簿"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "数据表"
        
        # 设置标题样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='1abc9c', end_color='1abc9c', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加标题行（不合并单元格，避免只读错误）
        title_cell = self.ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        
        # 设置行高
        self.ws.row_dimensions[1].height = 30
        
        return self.wb
    
    def write_data_with_header(self, data: List[Tuple], headers: List[str], start_row: int = 2):
        """写入带表头的数据"""
        if not self.ws:
            self.create_workbook()
        
        # 写入表头
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=start_row - 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # 获取表头和数据的最大长度
            header_val = len(str(headers[col_idx - 1]))
            data_vals = [len(str(self.ws.cell(row=r, column=col_idx).value or "")) 
                        for r in range(start_row, self.ws.max_row + 1)]
            
            max_length = max([header_val] + data_vals)
            self.ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        return True
    
    def save_workbook(self, filename: str):
        """保存工作簿"""
        if not self.wb:
            raise ValueError("没有活动的工作簿")
        
        # 确保目录存在
        directory = os.path.dirname(filename)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
        
        self.wb.save(filename)
        print(f"[Excel] 保存文件成功：{filename}")
        return True
    
    def read_workbook(self, filename: str, sheet_name: str = None):
        """读取工作簿"""
        if not os.path.exists(filename):
            raise FileNotFoundError(f"文件不存在：{filename}")
        
        self.wb = openpyxl.load_workbook(filename)
        
        if sheet_name:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.active
        
        print(f"[Excel] 读取文件成功：{filename}")
        return self.ws
    
    def get_all_data(self, min_row: int = 1, min_col: int = 1):
        """获取所有数据"""
        if not self.ws:
            raise ValueError("没有活动的工作表")
        
        data = []
        for row in self.ws.iter_rows(min_row=min_row, min_col=min_col, values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        
        return data
    
    def get_data_as_dict(self, headers: List[str], min_row: int = 2):
        """将数据转换为字典列表"""
        if not self.ws:
            raise ValueError("没有活动的工作表")
        
        data = []
        for row in self.ws.iter_rows(min_row=min_row, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]
                    else:
                        row_dict[header] = None
                data.append(row_dict)
        
        return data
    
    def export_income_records(self, records: List[Tuple], filename: str = None):
        """导出收入记录到 Excel"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data/exports/收入记录_{timestamp}.xlsx"
        
        headers = ["单据号", "日期", "收入类型", "金额", "支付方式", "备注"]
        
        self.create_workbook("收入记录导出")
        self.write_data_with_header(records, headers)
        self.save_workbook(filename)
        
        return filename
    
    def export_expense_records(self, records: List[Tuple], filename: str = None):
        """导出支出记录到 Excel"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data/exports/支出记录_{timestamp}.xlsx"
        
        headers = ["单据号", "日期", "支出类型", "金额", "支付方式", "备注"]
        
        self.create_workbook("支出记录导出")
        self.write_data_with_header(records, headers)
        self.save_workbook(filename)
        
        return filename
    
    def export_cash_flow(self, records: List[Tuple], filename: str = None):
        """导出流水账到 Excel"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data/exports/收支流水账_{timestamp}.xlsx"
        
        headers = ["日期", "序号", "收支类型", "单据号", "收入类型", "收入金额", 
                   "支出类型", "支出金额", "余额", "支付方式", "备注"]
        
        self.create_workbook("收支流水账导出")
        self.write_data_with_header(records, headers)
        self.save_workbook(filename)
        
        return filename
    
    def export_monthly_report(self, records: List[Tuple], year_month: str, filename: str = None):
        """导出月报表到 Excel"""
        if filename is None:
            filename = f"data/exports/月报表_{year_month}.xlsx"
        
        headers = ["账套号", "期初日期", "期末日期", "支付编码", 
                   "期初余额", "收入金额", "支出金额", "期末余额"]
        
        self.create_workbook(f"月报表_{year_month}")
        self.write_data_with_header(records, headers)
        self.save_workbook(filename)
        
        return filename


# 全局 Excel 处理器实例
excel_handler = ExcelHandler()
